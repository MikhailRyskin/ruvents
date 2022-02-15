from openpyxl import load_workbook
from sympy import isprime
from datetime import datetime, timedelta

wb = load_workbook('task_support.xlsx')
sheet = wb['Tasks']

num1_count = 0
num2_count = 0
num3_count = 0
date1_count = 0
date2_count = 0
date3_count = 0

for row in range(3, 1003):
    cell_1 = sheet.cell(row=row, column=2)
    if int(cell_1.value) % 2 == 0:
        num1_count += 1

    cell_2 = sheet.cell(row=row, column=3)
    if isprime(int(cell_2.value)):
        num2_count += 1

    cell_3 = sheet.cell(row=row, column=4)
    cell_3_value = float(cell_3.value.replace(' ', '').replace(',', '.'))
    if cell_3_value < 0.5:
        num3_count += 1

    cell_4 = sheet.cell(row=row, column=5)
    if cell_4.value.startswith('Tue'):
        date1_count += 1

    cell_5 = sheet.cell(row=row, column=6)
    cell_5_day_number = datetime.strptime(cell_5.value[:10], '%Y-%m-%d').weekday()
    if cell_5_day_number == 1:
        date2_count += 1

    cell_6 = sheet.cell(row=row, column=7)
    cell_6_date = datetime.strptime(cell_6.value, '%m-%d-%Y')
    if cell_6_date.weekday() == 1:
        week_delta = timedelta(days=7)
        next_tuesday = cell_6_date + week_delta
        if cell_6_date.month != next_tuesday.month:
            date3_count += 1


print(num1_count, num2_count, num3_count, date1_count, date2_count, date3_count)
